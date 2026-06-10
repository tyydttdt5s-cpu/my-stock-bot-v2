import yfinance as yf
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ==========================
# 持股清單
# ==========================

my_stocks = [
    "0052",
    "2303",
    "3711"
]

# ==========================
# 從CSV讀取股票名單
# ==========================

stock_df = pd.read_csv(
    "股票名單.csv",
    encoding="utf-8-sig"
)

stocks = dict(
    zip(
        stock_df["股票名稱"],
        stock_df["股票代號"]
    )
)

results = []

print("開始分析股票...")

for name, ticker in stocks.items():

    try:

        df = yf.download(
            ticker,
            period="6mo",
            progress=False,
            auto_adjust=True
        )

        if df.empty:
            print(f"{name} 無資料")
            continue

        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)

        close = df["Close"].dropna()

        if len(close) < 30:
            continue

        current_price = float(close.iloc[-1])

        ma5 = float(
            close.rolling(5).mean().iloc[-1]
        )

        ma20 = float(
            close.rolling(20).mean().iloc[-1]
        )

        # RSI計算

        delta = close.diff()

        gain = delta.where(
            delta > 0,
            0
        )

        loss = -delta.where(
            delta < 0,
            0
        )

        avg_gain = gain.rolling(14).mean()

        avg_loss = loss.rolling(14).mean()

        rs = avg_gain / avg_loss

        rsi = float(
            (
                100
                - (
                    100
                    / (1 + rs)
                )
            ).iloc[-1]
        )

        # RSI狀態

        if rsi >= 70:

            rsi_status = "過熱"

        elif rsi >= 60:

            rsi_status = "偏強"

        elif rsi >= 40:

            rsi_status = "中性"

        elif rsi >= 30:

            rsi_status = "偏弱"

        else:

            rsi_status = "超跌"

        # 均線趨勢

        if current_price > ma5 > ma20:

            trend = "多頭排列"

        elif current_price < ma5 < ma20:

            trend = "空頭排列"

        else:

            trend = "盤整"

        # 今日漲跌%

        change_pct = (
            (close.iloc[-1] - close.iloc[-2])
            / close.iloc[-2]
            * 100
        )

        # 20日乖離率%

        bias = (
            (current_price - ma20)
            / ma20
            * 100
        )

        # 技術評分

        score = 0

        if current_price > ma20:
            score += 30

        if ma5 > ma20:
            score += 30

        if 50 <= rsi <= 70:
            score += 20

        if change_pct > 0:
            score += 10

        if abs(bias) <= 15:
            score += 10

        # AI建議

        if score >= 90:

            advice = "🔥強勢續抱"

        elif score >= 80:

            advice = "✅可觀察"

        elif score >= 60:

            advice = "👀觀察"

        elif score >= 40:

            advice = "⚠️偏弱"

        else:

            advice = "❌避開"

        stock_code = ticker.replace(
            ".TW",
            ""
        )

        hold = (
            "✓"
            if stock_code in my_stocks
            else ""
        )

        results.append([
            hold,
            name,
            round(current_price, 2),
            round(ma5, 2),
            round(ma20, 2),
            trend,
            round(rsi, 2),
            rsi_status,
            round(bias, 2),
            round(change_pct, 2),
            score,
            advice
        ])

        print(f"{name} 完成")

    except Exception as e:

        print(
            f"{name} 發生錯誤：{e}"
        )

columns = [
    "持股",
    "股票",
    "現價",
    "5日均線",
    "20日均線",
    "均線趨勢",
    "RSI強弱指標",
    "RSI狀態",
    "20日乖離率%",
    "今日漲跌%",
    "技術評分",
    "AI建議"
]

result_df = pd.DataFrame(
    results,
    columns=columns
)

result_df = result_df.sort_values(
    by="技術評分",
    ascending=False
)

file_name = "股票分析V8.xlsx"

result_df.to_excel(
    file_name,
    index=False
)

# ==========================
# Excel格式化
# ==========================

wb = load_workbook(file_name)

ws = wb.active

green_fill = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid"
)

yellow_fill = PatternFill(
    start_color="FFF2CC",
    end_color="FFF2CC",
    fill_type="solid"
)

red_fill = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)

score_col = None

for cell in ws[1]:

    if cell.value == "技術評分":

        score_col = cell.column

        break

for row in range(
    2,
    ws.max_row + 1
):

    score = ws.cell(
        row,
        score_col
    ).value

    if score >= 80:

        ws.cell(
            row,
            score_col
        ).fill = green_fill

    elif score >= 60:

        ws.cell(
            row,
            score_col
        ).fill = yellow_fill

    else:

        ws.cell(
            row,
            score_col
        ).fill = red_fill

for column in ws.columns:

    max_length = 0

    column_letter = get_column_letter(
        column[0].column
    )

    for cell in column:

        try:

            if len(str(cell.value)) > max_length:

                max_length = len(
                    str(cell.value)
                )

        except:
            pass

    ws.column_dimensions[
        column_letter
    ].width = max_length + 4

wb.save(file_name)

print("\n====================")
print("V8 分析完成")
print(f"輸出檔案：{file_name}")
print("====================")